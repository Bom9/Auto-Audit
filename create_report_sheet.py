from openpyxl import Workbook
from openpyxl.styles import PatternFill
import argparse
import json
import glob
import os
from sys import platform

def main(ARGS):
    wb = Workbook()
    ws = wb.active

    for f in glob.glob(f"{os.path.abspath(os.curdir)}/*.json", recursive=True):
        if platform == "linux":
            filename = f.split('/')[-1]
        elif platform == "win32":
            filename = f.split('\\')[-1]
        hostname = filename.split('_')[0]
        if ws.title == 'Sheet':
            ws.title = hostname
        else:
            ws = wb.create_sheet(hostname)
        print(f"Opening {filename} ...")
        with open(f) as results_file:
            data = json.loads(results_file.read())
        print(f"Converting {filename}")
        for n in range(0, len(data['results'])):
            title = data['results'][n]['title']
            if title == 'Benchmark MetaData':
                continue
            id = data['results'][n]['meta']['CIS_ID']
            successful = data['results'][n]['successful']
            if isinstance(id, str):
                ws.cell(row=n+1, column=1, value=id)
            elif isinstance(id, list):
                new_id = str(id)
                ws.cell(row=n+1, column=1, value=new_id)
            if successful:
                c = ws.cell(row=n+1, column=2, value=successful)
                c.fill = PatternFill("solid", fgColor="0000FF00")
            else:
                c = ws.cell(row=n+1, column=2, value=successful)
                c.fill = PatternFill("solid", fgColor="00FF0000")
            ws.cell(row=n+1, column=3, value=title)

    wb.save(ARGS.filename)
    print("Conversion completed...")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", help="The filename for the spreadsheet that will be output")
    args = parser.parse_args()
    main(args)
